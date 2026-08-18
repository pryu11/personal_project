"""Filter raw LCA disclosure data down to California tech/data roles.

Streams rows directly from each raw Excel file instead of loading it into a
pandas DataFrame first -- a fiscal quarter file has hundreds of thousands of
rows, but only a small fraction are both in California and in a target SOC
code, so filtering row-by-row keeps memory usage low regardless of file size.

Raw files are discovered by scanning data/raw/ recursively and identified by
the FY<year>_Q<quarter> pattern in their filename, not by which folder they
happen to sit in -- that way a misplaced or newly-added file still gets
picked up instead of silently skipped.
"""

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from reference.tech_soc_codes import TECH_SOC_CODES

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TARGET_STATE = "CA"
TARGET_SOC_PREFIXES = tuple(TECH_SOC_CODES.keys())
FISCAL_LABEL_RE = re.compile(r"FY(\d{4})_Q(\d)", re.IGNORECASE)


def extract_filter(raw_path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(raw_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    col_index = {name: i for i, name in enumerate(header)}
    state_i = col_index["WORKSITE_STATE"]
    soc_i = col_index["SOC_CODE"]

    matched = []
    total = 0
    for row in rows:
        total += 1
        state = row[state_i]
        soc = row[soc_i]
        if state != TARGET_STATE:
            continue
        if not soc or not soc.startswith(TARGET_SOC_PREFIXES):
            continue
        matched.append(row)

    wb.close()

    print(f"{raw_path.name}: {total:,} total rows -> {len(matched):,} rows after CA + tech SOC filter")
    return pd.DataFrame(matched, columns=header)


def fiscal_label(raw_path: Path) -> str:
    match = FISCAL_LABEL_RE.search(raw_path.stem)
    if not match:
        raise ValueError(f"Can't determine fiscal year/quarter from filename: {raw_path.name}")
    year, quarter = match.groups()
    return f"FY{year}_Q{quarter}"


def main():
    raw_dir = PROJECT_ROOT / "data" / "raw"
    raw_files = sorted(raw_dir.glob("**/*.xlsx"))
    if not raw_files:
        raise FileNotFoundError(f"No .xlsx files found under {raw_dir}")

    out_dir = PROJECT_ROOT / "data" / "interim"
    out_dir.mkdir(parents=True, exist_ok=True)

    for raw_path in raw_files:
        label = fiscal_label(raw_path)
        out_path = out_dir / f"lca_{label}_ca_tech.csv"
        df = extract_filter(raw_path)
        df.to_csv(out_path, index=False)
        print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
